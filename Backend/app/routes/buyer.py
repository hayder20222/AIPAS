from fastapi import APIRouter, Depends, HTTPException, status, UploadFile, File
from fastapi.responses import StreamingResponse
from typing import List
from bson import ObjectId
from datetime import datetime
from io import BytesIO
from app.models.request import ProcurementRequestCreate, ProcurementRequestResponse
from app.models.quotation import QuotationResponse
from app.database import get_database
from app.routes.auth import get_current_user
from app.utils.helpers import serialize_objectid
from app.services.llama_service import llama_service
from app.services.pdf_processor import PDFProcessor

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

router = APIRouter(prefix="/buyer", tags=["Buyer"])

def verify_buyer(current_user: dict = Depends(get_current_user)):
    if current_user["role"] != "buyer":
        raise HTTPException(status_code=403, detail="Only buyers can access this endpoint")
    return current_user

@router.post("/requests", response_model=ProcurementRequestResponse)
async def create_request(
    request: ProcurementRequestCreate,
    current_user: dict = Depends(verify_buyer)
):
    db = await get_database()
    
    request_dict = request.model_dump()
    request_dict["buyer_id"] = current_user["_id"]
    request_dict["status"] = "open"
    request_dict["quotations_received"] = 0
    request_dict["posted_at"] = datetime.utcnow()
    request_dict["updated_at"] = datetime.utcnow()
    
    result = await db.procurement_requests.insert_one(request_dict)
    
    created_request = await db.procurement_requests.find_one({"_id": result.inserted_id})
    created_request = serialize_objectid(created_request)
    
    return ProcurementRequestResponse(
        id=created_request["_id"],
        buyer_id=str(created_request["buyer_id"]),
        title=created_request["title"],
        description=created_request.get("description"),
        items_needed=created_request["items_needed"],
        budget=created_request["budget"],
        deadline=created_request["deadline"],
        status=created_request["status"],
        quotations_received=created_request["quotations_received"],
        posted_at=created_request["posted_at"]
    )

@router.get("/requests", response_model=List[ProcurementRequestResponse])
async def get_my_requests(current_user: dict = Depends(verify_buyer)):
    db = await get_database()
    
    cursor = db.procurement_requests.find({"buyer_id": current_user["_id"]}).sort("posted_at", -1)
    requests = await cursor.to_list(length=100)
    
    return [
        ProcurementRequestResponse(
            id=str(req["_id"]),
            buyer_id=str(req["buyer_id"]),
            title=req["title"],
            description=req.get("description"),
            items_needed=req["items_needed"],
            budget=req["budget"],
            deadline=req["deadline"],
            status=req["status"],
            quotations_received=req["quotations_received"],
            posted_at=req["posted_at"]
        )
        for req in requests
    ]

@router.get("/requests/{request_id}", response_model=ProcurementRequestResponse)
async def get_request_detail(
    request_id: str,
    current_user: dict = Depends(verify_buyer)
):
    db = await get_database()
    
    request = await db.procurement_requests.find_one({
        "_id": ObjectId(request_id),
        "buyer_id": current_user["_id"]
    })
    
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    
    request = serialize_objectid(request)
    
    return ProcurementRequestResponse(
        id=request["_id"],
        buyer_id=str(request["buyer_id"]),
        title=request["title"],
        description=request.get("description"),
        items_needed=request["items_needed"],
        budget=request["budget"],
        deadline=request["deadline"],
        status=request["status"],
        quotations_received=request["quotations_received"],
        posted_at=request["posted_at"]
    )

@router.get("/requests/{request_id}/quotations", response_model=List[QuotationResponse])
async def get_quotations_for_request(
    request_id: str,
    current_user: dict = Depends(verify_buyer)
):
    db = await get_database()
    
    # Verify request belongs to buyer
    request = await db.procurement_requests.find_one({
        "_id": ObjectId(request_id),
        "buyer_id": current_user["_id"]
    })
    
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    
    # Get quotations
    cursor = db.quotations.find({"request_id": request_id}).sort("total_amount", 1)
    quotations = await cursor.to_list(length=100)
    
    return [
        QuotationResponse(
            id=str(quot["_id"]),
            request_id=quot["request_id"],
            vendor_id=str(quot["vendor_id"]),
            vendor_name=quot["vendor_name"],
            items=quot["items"],
            delivery_time_days=quot.get("delivery_time_days"),
            payment_terms=quot.get("payment_terms"),
            valid_until=quot.get("valid_until"),
            notes=quot.get("notes"),
            total_amount=quot["total_amount"],
            status=quot["status"],
            submitted_at=quot["submitted_at"],
            pdf_filename=quot["pdf_file"]["filename"]
        )
        for quot in quotations
    ]

@router.post("/requests/{request_id}/quotations/{quotation_id}/accept")
async def accept_quotation(
    request_id: str,
    quotation_id: str,
    current_user: dict = Depends(verify_buyer)
):
    db = await get_database()
    
    # Verify request belongs to buyer
    request = await db.procurement_requests.find_one({
        "_id": ObjectId(request_id),
        "buyer_id": current_user["_id"]
    })
    
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    
    # Update quotation status
    result = await db.quotations.update_one(
        {"_id": ObjectId(quotation_id), "request_id": request_id},
        {"$set": {"status": "accepted"}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Quotation not found")
    
    # Update request status
    await db.procurement_requests.update_one(
        {"_id": ObjectId(request_id)},
        {"$set": {"status": "awarded", "updated_at": datetime.utcnow()}}
    )
    
    return {"message": "Quotation accepted successfully"}

@router.get("/requests/{request_id}/comparison")
async def get_comparison_summary(
    request_id: str,
    current_user: dict = Depends(verify_buyer)
):
    db = await get_database()
    
    # Verify request belongs to buyer
    request = await db.procurement_requests.find_one({
        "_id": ObjectId(request_id),
        "buyer_id": current_user["_id"]
    })
    
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    
    # Get quotations
    cursor = db.quotations.find({"request_id": request_id})
    quotations = await cursor.to_list(length=100)
    
    if not quotations:
        return {"summary": "No quotations received yet"}
    
    # Prepare data for AI
    comparison_data = [
        {
            "vendor": quot["vendor_name"],
            "total_amount": quot["total_amount"],
            "delivery_days": quot.get("delivery_time_days"),
            "items_count": len(quot["items"])
        }
        for quot in quotations
    ]
    
    # Generate AI summary
    summary = await llama_service.compare_quotations(comparison_data)
    
    return {
        "summary": summary,
        "quotations_count": len(quotations),
        "lowest_price": min(q["total_amount"] for q in quotations),
        "highest_price": max(q["total_amount"] for q in quotations)
    }

@router.get("/requests/{request_id}/export-excel")
async def export_quotations_excel(
    request_id: str,
    current_user: dict = Depends(verify_buyer)
):
    """Export all quotations for a request to Excel with comparison analysis"""
    
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Excel export not available. Install openpyxl.")
    
    db = await get_database()
    
    # Verify request belongs to buyer
    request = await db.procurement_requests.find_one({
        "_id": ObjectId(request_id),
        "buyer_id": current_user["_id"]
    })
    
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    
    # Get quotations
    cursor = db.quotations.find({"request_id": request_id}).sort("total_amount", 1)
    quotations = await cursor.to_list(length=100)
    
    if not quotations:
        raise HTTPException(status_code=404, detail="No quotations found for this request")
    
    # Create workbook
    wb = Workbook()
    
    # === Sheet 1: Summary Comparison ===
    ws_summary = wb.active
    ws_summary.title = "Summary Comparison"
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
    highlight_fill = PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid")  # Green for lowest
    warning_fill = PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid")  # Red for highest
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Calculate statistics
    prices = [q["total_amount"] for q in quotations]
    lowest_price = min(prices)
    highest_price = max(prices)
    avg_price = sum(prices) / len(prices) if prices else 0
    potential_savings = highest_price - lowest_price
    
    # Title
    ws_summary['A1'] = f"Quotation Comparison Report - {request['title']}"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary.merge_cells('A1:G1')
    
    # Request Info
    ws_summary['A3'] = "Request Details"
    ws_summary['A3'].font = Font(bold=True, size=12)
    ws_summary['A4'] = f"Title: {request['title']}"
    ws_summary['A5'] = f"Budget: ${request.get('budget', 'N/A'):,.2f}" if request.get('budget') else "Budget: N/A"
    ws_summary['A6'] = f"Deadline: {request.get('deadline', 'N/A')}"
    ws_summary['A7'] = f"Total Quotations: {len(quotations)}"
    
    # Statistics Section
    ws_summary['A9'] = "Price Analysis"
    ws_summary['A9'].font = Font(bold=True, size=12)
    
    stats_data = [
        ["Metric", "Value"],
        ["Lowest Price", f"${lowest_price:,.2f}"],
        ["Highest Price", f"${highest_price:,.2f}"],
        ["Average Price", f"${avg_price:,.2f}"],
        ["Potential Savings", f"${potential_savings:,.2f}"],
        ["Price Range %", f"{((highest_price - lowest_price) / lowest_price * 100):.1f}%" if lowest_price > 0 else "N/A"]
    ]
    
    for row_idx, row_data in enumerate(stats_data, start=10):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 10:
                cell.font = header_font
                cell.fill = header_fill
    
    # Vendor Comparison Table
    ws_summary['A18'] = "Vendor Comparison"
    ws_summary['A18'].font = Font(bold=True, size=12)
    
    comparison_headers = ["Rank", "Vendor Name", "Total Amount", "Delivery Days", "Items Count", "Price vs Lowest", "Status"]
    for col_idx, header in enumerate(comparison_headers, start=1):
        cell = ws_summary.cell(row=19, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center')
    
    for row_idx, quot in enumerate(quotations, start=20):
        price_diff = ((quot["total_amount"] - lowest_price) / lowest_price * 100) if lowest_price > 0 else 0
        
        row_data = [
            row_idx - 19,  # Rank
            quot["vendor_name"],
            quot["total_amount"],
            quot.get("delivery_time_days", "N/A"),
            len(quot["items"]),
            f"+{price_diff:.1f}%" if price_diff > 0 else "Lowest",
            quot["status"].upper()
        ]
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            # Format price column
            if col_idx == 3:
                cell.number_format = '$#,##0.00'
            
            # Highlight lowest price row
            if quot["total_amount"] == lowest_price:
                cell.fill = highlight_fill
            elif quot["total_amount"] == highest_price:
                cell.fill = warning_fill
    
    # Adjust column widths
    for col in range(1, 8):
        ws_summary.column_dimensions[get_column_letter(col)].width = 18
    
    # === Sheet 2: Detailed Items ===
    ws_items = wb.create_sheet("Item Details")
    
    item_headers = ["Vendor", "Item Description", "Quantity", "Unit Price", "Total Price", "Brand", "Warranty"]
    for col_idx, header in enumerate(item_headers, start=1):
        cell = ws_items.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    row_idx = 2
    for quot in quotations:
        for item in quot["items"]:
            ws_items.cell(row=row_idx, column=1, value=quot["vendor_name"]).border = thin_border
            ws_items.cell(row=row_idx, column=2, value=item.get("product_name", "N/A")).border = thin_border
            ws_items.cell(row=row_idx, column=3, value=item.get("quantity", 1)).border = thin_border
            
            unit_price_cell = ws_items.cell(row=row_idx, column=4, value=item.get("unit_price", 0))
            unit_price_cell.number_format = '$#,##0.00'
            unit_price_cell.border = thin_border
            
            total_price_cell = ws_items.cell(row=row_idx, column=5, value=item.get("total_price", 0))
            total_price_cell.number_format = '$#,##0.00'
            total_price_cell.border = thin_border
            
            ws_items.cell(row=row_idx, column=6, value=item.get("brand", "N/A")).border = thin_border
            ws_items.cell(row=row_idx, column=7, value=item.get("warranty", "N/A")).border = thin_border
            row_idx += 1
    
    # Adjust column widths
    ws_items.column_dimensions['A'].width = 25
    ws_items.column_dimensions['B'].width = 40
    ws_items.column_dimensions['C'].width = 12
    ws_items.column_dimensions['D'].width = 15
    ws_items.column_dimensions['E'].width = 15
    ws_items.column_dimensions['F'].width = 15
    ws_items.column_dimensions['G'].width = 15
    
    # === Sheet 3: Price Breakdown by Item ===
    ws_breakdown = wb.create_sheet("Price Breakdown")
    
    # Get all unique items across quotations
    all_items = {}
    for quot in quotations:
        for item in quot["items"]:
            item_name = item.get("product_name", "Unknown")
            if item_name not in all_items:
                all_items[item_name] = {}
            all_items[item_name][quot["vendor_name"]] = item.get("unit_price", 0)
    
    vendor_names = [q["vendor_name"] for q in quotations]
    
    # Header row
    breakdown_headers = ["Item"] + vendor_names + ["Lowest", "Highest", "Savings"]
    for col_idx, header in enumerate(breakdown_headers, start=1):
        cell = ws_breakdown.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    row_idx = 2
    for item_name, vendor_prices in all_items.items():
        ws_breakdown.cell(row=row_idx, column=1, value=item_name).border = thin_border
        
        prices_list = []
        for col_idx, vendor in enumerate(vendor_names, start=2):
            price = vendor_prices.get(vendor, 0)
            cell = ws_breakdown.cell(row=row_idx, column=col_idx, value=price if price > 0 else "N/A")
            cell.border = thin_border
            if price > 0:
                cell.number_format = '$#,##0.00'
                prices_list.append(price)
        
        # Add min, max, savings
        if prices_list:
            min_price = min(prices_list)
            max_price = max(prices_list)
            savings = max_price - min_price
            
            min_cell = ws_breakdown.cell(row=row_idx, column=len(vendor_names) + 2, value=min_price)
            min_cell.number_format = '$#,##0.00'
            min_cell.fill = highlight_fill
            min_cell.border = thin_border
            
            max_cell = ws_breakdown.cell(row=row_idx, column=len(vendor_names) + 3, value=max_price)
            max_cell.number_format = '$#,##0.00'
            max_cell.fill = warning_fill
            max_cell.border = thin_border
            
            savings_cell = ws_breakdown.cell(row=row_idx, column=len(vendor_names) + 4, value=savings)
            savings_cell.number_format = '$#,##0.00'
            savings_cell.border = thin_border
        
        row_idx += 1
    
    # Adjust column widths
    ws_breakdown.column_dimensions['A'].width = 35
    for col in range(2, len(breakdown_headers) + 1):
        ws_breakdown.column_dimensions[get_column_letter(col)].width = 15
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Generate filename
    filename = f"quotation_comparison_{request['title'][:30].replace(' ', '_')}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


# ==================== QUICK COMPARE FEATURE ====================
# Upload and compare PDFs received from WhatsApp, Email, etc.

@router.post("/quick-compare")
async def quick_compare_pdfs(
    pdfs: List[UploadFile] = File(...),
    current_user: dict = Depends(verify_buyer)
):
    """
    Upload multiple PDF quotations and get instant comparison.
    Useful for comparing quotations received via WhatsApp, Email, etc.
    """
    
    if len(pdfs) < 1:
        raise HTTPException(status_code=400, detail="Please upload at least 1 PDF")
    
    if len(pdfs) > 10:
        raise HTTPException(status_code=400, detail="Maximum 10 PDFs allowed at once")
    
    parsed_quotations = []
    
    for idx, pdf_file in enumerate(pdfs):
        if not pdf_file.filename.lower().endswith('.pdf'):
            continue
        
        try:
            # Read PDF content
            pdf_bytes = await pdf_file.read()
            
            # Use Vision API for accurate extraction
            parsed_data = await PDFProcessor.process_pdf_complete(pdf_bytes)
            
            # Use filename as vendor name if not detected
            vendor_name = parsed_data.get("vendor_name", "Unknown Vendor")
            if vendor_name == "Unknown Vendor":
                # Clean filename to use as vendor name
                vendor_name = pdf_file.filename.replace('.pdf', '').replace('_', ' ').replace('-', ' ').title()
            
            quotation = {
                "id": f"quick_{idx + 1}",
                "filename": pdf_file.filename,
                "vendor_name": vendor_name,
                "items": parsed_data.get("items", []),
                "total_amount": parsed_data.get("total_amount", 0),
                "subtotal": parsed_data.get("subtotal"),
                "tax_amount": parsed_data.get("tax_amount"),
                "delivery_time_days": parsed_data.get("delivery_time_days"),
                "payment_terms": parsed_data.get("payment_terms"),
                "dates_found": parsed_data.get("dates_found", []),
                "prices_found": parsed_data.get("prices_found", [])[:5],
                "extraction_success": len(parsed_data.get("items", [])) > 0 or parsed_data.get("total_amount", 0) > 0
            }
            
            parsed_quotations.append(quotation)
            
        except Exception as e:
            parsed_quotations.append({
                "id": f"quick_{idx + 1}",
                "filename": pdf_file.filename,
                "vendor_name": pdf_file.filename.replace('.pdf', ''),
                "items": [],
                "total_amount": 0,
                "error": str(e),
                "extraction_success": False
            })
    
    if not parsed_quotations:
        raise HTTPException(status_code=400, detail="No valid PDFs could be processed")
    
    # Calculate comparison stats
    successful = [q for q in parsed_quotations if q.get("extraction_success")]
    prices = [q["total_amount"] for q in successful if q["total_amount"] > 0]
    
    comparison_stats = {
        "total_uploaded": len(pdfs),
        "successfully_parsed": len(successful),
        "failed": len(parsed_quotations) - len(successful),
        "lowest_price": min(prices) if prices else 0,
        "highest_price": max(prices) if prices else 0,
        "average_price": sum(prices) / len(prices) if prices else 0,
        "potential_savings": (max(prices) - min(prices)) if len(prices) >= 2 else 0
    }
    
    return {
        "quotations": parsed_quotations,
        "comparison": comparison_stats
    }


@router.post("/quick-compare/export-excel")
async def export_quick_compare_excel(
    pdfs: List[UploadFile] = File(...),
    current_user: dict = Depends(verify_buyer)
):
    """
    Upload multiple PDFs and directly get Excel comparison file.
    """
    
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Excel export not available")
    
    if len(pdfs) < 1:
        raise HTTPException(status_code=400, detail="Please upload at least 1 PDF")
    
    # Parse all PDFs
    parsed_quotations = []
    
    for idx, pdf_file in enumerate(pdfs):
        if not pdf_file.filename.lower().endswith('.pdf'):
            continue
        
        try:
            pdf_bytes = await pdf_file.read()
            parsed_data = await PDFProcessor.process_pdf_complete(pdf_bytes)
            
            vendor_name = parsed_data.get("vendor_name", "Unknown Vendor")
            if vendor_name == "Unknown Vendor":
                vendor_name = pdf_file.filename.replace('.pdf', '').replace('_', ' ').replace('-', ' ').title()
            
            parsed_quotations.append({
                "vendor_name": vendor_name,
                "filename": pdf_file.filename,
                "items": parsed_data.get("items", []),
                "total_amount": parsed_data.get("total_amount", 0),
                "subtotal": parsed_data.get("subtotal"),
                "tax_amount": parsed_data.get("tax_amount"),
                "delivery_time_days": parsed_data.get("delivery_time_days"),
                "payment_terms": parsed_data.get("payment_terms"),
            })
        except Exception as e:
            parsed_quotations.append({
                "vendor_name": pdf_file.filename.replace('.pdf', ''),
                "filename": pdf_file.filename,
                "items": [],
                "total_amount": 0,
                "error": str(e)
            })
    
    if not parsed_quotations:
        raise HTTPException(status_code=400, detail="No valid PDFs could be processed")
    
    # Create Excel workbook
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
    highlight_fill = PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid")
    warning_fill = PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid")
    info_fill = PatternFill(start_color="BEE3F8", end_color="BEE3F8", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Calculate statistics
    prices = [q["total_amount"] for q in parsed_quotations if q["total_amount"] > 0]
    lowest_price = min(prices) if prices else 0
    highest_price = max(prices) if prices else 0
    avg_price = sum(prices) / len(prices) if prices else 0
    
    # === Sheet 1: Summary ===
    ws_summary = wb.active
    ws_summary.title = "Summary Comparison"
    
    # Title
    ws_summary['A1'] = "Quick Quotation Comparison Report"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary.merge_cells('A1:G1')
    
    ws_summary['A2'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC"
    ws_summary['A2'].font = Font(italic=True, color="666666")
    
    # Statistics
    ws_summary['A4'] = "Price Analysis"
    ws_summary['A4'].font = Font(bold=True, size=12)
    
    stats = [
        ["Metric", "Value"],
        ["PDFs Analyzed", len(parsed_quotations)],
        ["Lowest Price", f"${lowest_price:,.2f}" if lowest_price else "N/A"],
        ["Highest Price", f"${highest_price:,.2f}" if highest_price else "N/A"],
        ["Average Price", f"${avg_price:,.2f}" if avg_price else "N/A"],
        ["Potential Savings", f"${highest_price - lowest_price:,.2f}" if len(prices) >= 2 else "N/A"],
    ]
    
    for row_idx, row_data in enumerate(stats, start=5):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 5:
                cell.font = header_font
                cell.fill = header_fill
    
    # Vendor Comparison
    ws_summary['A13'] = "Vendor Comparison"
    ws_summary['A13'].font = Font(bold=True, size=12)
    
    headers = ["Rank", "Vendor/File", "Total Amount", "Items Count", "Delivery Days", "Subtotal", "Tax", "Price vs Lowest"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_summary.cell(row=14, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    # Sort by total amount
    sorted_quotations = sorted(parsed_quotations, key=lambda x: x["total_amount"] if x["total_amount"] > 0 else float('inf'))
    
    for row_idx, quot in enumerate(sorted_quotations, start=15):
        price_diff = ((quot["total_amount"] - lowest_price) / lowest_price * 100) if lowest_price > 0 and quot["total_amount"] > 0 else 0
        
        row_data = [
            row_idx - 14,
            quot["vendor_name"],
            quot["total_amount"],
            len(quot["items"]),
            quot.get("delivery_time_days") or "N/A",
            quot.get("subtotal") or "N/A",
            quot.get("tax_amount") or "N/A",
            f"+{price_diff:.1f}%" if price_diff > 0 else ("Lowest" if quot["total_amount"] == lowest_price and lowest_price > 0 else "N/A")
        ]
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            if col_idx == 3 and isinstance(value, (int, float)):
                cell.number_format = '$#,##0.00'
            
            # Highlight rows
            if quot["total_amount"] == lowest_price and lowest_price > 0:
                cell.fill = highlight_fill
            elif quot["total_amount"] == highest_price and highest_price > 0 and len(prices) > 1:
                cell.fill = warning_fill
    
    # Adjust columns
    for col in range(1, 9):
        ws_summary.column_dimensions[get_column_letter(col)].width = 18
    
    # === Sheet 2: All Items ===
    ws_items = wb.create_sheet("All Items")
    
    item_headers = ["Vendor", "Item Description", "Quantity", "Unit Price", "Total Price"]
    for col_idx, header in enumerate(item_headers, start=1):
        cell = ws_items.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    row_idx = 2
    for quot in sorted_quotations:
        for item in quot["items"]:
            ws_items.cell(row=row_idx, column=1, value=quot["vendor_name"]).border = thin_border
            ws_items.cell(row=row_idx, column=2, value=item.get("product_name", "N/A")).border = thin_border
            ws_items.cell(row=row_idx, column=3, value=item.get("quantity", 1)).border = thin_border
            
            unit_cell = ws_items.cell(row=row_idx, column=4, value=item.get("unit_price", 0))
            unit_cell.number_format = '$#,##0.00'
            unit_cell.border = thin_border
            
            total_cell = ws_items.cell(row=row_idx, column=5, value=item.get("total_price", 0))
            total_cell.number_format = '$#,##0.00'
            total_cell.border = thin_border
            
            row_idx += 1
    
    ws_items.column_dimensions['A'].width = 25
    ws_items.column_dimensions['B'].width = 45
    ws_items.column_dimensions['C'].width = 12
    ws_items.column_dimensions['D'].width = 15
    ws_items.column_dimensions['E'].width = 15
    
    # === Sheet 3: Item-wise Price Comparison ===
    ws_compare = wb.create_sheet("Item Price Comparison")
    
    # Gather all unique items
    all_items = {}
    for quot in sorted_quotations:
        for item in quot["items"]:
            item_name = item.get("product_name", "Unknown")
            if item_name not in all_items:
                all_items[item_name] = {}
            all_items[item_name][quot["vendor_name"]] = item.get("unit_price", 0)
    
    vendor_names = [q["vendor_name"] for q in sorted_quotations]
    
    compare_headers = ["Item"] + vendor_names + ["Lowest", "Highest", "Savings"]
    for col_idx, header in enumerate(compare_headers, start=1):
        cell = ws_compare.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    row_idx = 2
    for item_name, vendor_prices in all_items.items():
        ws_compare.cell(row=row_idx, column=1, value=item_name).border = thin_border
        
        prices_list = []
        for col_idx, vendor in enumerate(vendor_names, start=2):
            price = vendor_prices.get(vendor, 0)
            cell = ws_compare.cell(row=row_idx, column=col_idx, value=price if price > 0 else "N/A")
            cell.border = thin_border
            if price > 0:
                cell.number_format = '$#,##0.00'
                prices_list.append(price)
        
        if prices_list:
            min_p = min(prices_list)
            max_p = max(prices_list)
            
            min_cell = ws_compare.cell(row=row_idx, column=len(vendor_names) + 2, value=min_p)
            min_cell.number_format = '$#,##0.00'
            min_cell.fill = highlight_fill
            min_cell.border = thin_border
            
            max_cell = ws_compare.cell(row=row_idx, column=len(vendor_names) + 3, value=max_p)
            max_cell.number_format = '$#,##0.00'
            max_cell.fill = warning_fill
            max_cell.border = thin_border
            
            savings_cell = ws_compare.cell(row=row_idx, column=len(vendor_names) + 4, value=max_p - min_p)
            savings_cell.number_format = '$#,##0.00'
            savings_cell.border = thin_border
        
        row_idx += 1
    
    ws_compare.column_dimensions['A'].width = 40
    for col in range(2, len(compare_headers) + 1):
        ws_compare.column_dimensions[get_column_letter(col)].width = 15
    
    # Save
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"quick_comparison_{datetime.utcnow().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@router.post("/requests/{request_id}/export-combined-excel")
async def export_combined_excel(
    request_id: str,
    pdfs: List[UploadFile] = File(default=[]),
    current_user: dict = Depends(verify_buyer)
):
    """
    Export both vendor-submitted quotations AND uploaded PDFs to a single Excel file.
    """
    
    if not EXCEL_AVAILABLE:
        raise HTTPException(status_code=500, detail="Excel export not available")
    
    db = await get_database()
    
    # Verify request belongs to buyer
    request = await db.procurement_requests.find_one({
        "_id": ObjectId(request_id),
        "buyer_id": current_user["_id"]
    })
    
    if not request:
        raise HTTPException(status_code=404, detail="Request not found")
    
    # Get vendor quotations from database
    cursor = db.quotations.find({"request_id": request_id}).sort("total_amount", 1)
    db_quotations = await cursor.to_list(length=100)
    
    all_quotations = []
    
    # Add database quotations
    for quot in db_quotations:
        all_quotations.append({
            "vendor_name": quot["vendor_name"],
            "source": "Vendor Submission",
            "items": quot["items"],
            "total_amount": quot["total_amount"],
            "delivery_time_days": quot.get("delivery_time_days"),
            "payment_terms": quot.get("payment_terms"),
            "status": quot["status"]
        })
    
    # Parse uploaded PDFs and add them
    for idx, pdf_file in enumerate(pdfs):
        if not pdf_file.filename.lower().endswith('.pdf'):
            continue
        
        try:
            pdf_bytes = await pdf_file.read()
            parsed_data = await PDFProcessor.process_pdf_complete(pdf_bytes)
            
            vendor_name = parsed_data.get("vendor_name", "Unknown Vendor")
            if vendor_name == "Unknown Vendor":
                vendor_name = pdf_file.filename.replace('.pdf', '').replace('_', ' ').title()
            
            all_quotations.append({
                "vendor_name": vendor_name,
                "source": f"Uploaded: {pdf_file.filename}",
                "items": parsed_data.get("items", []),
                "total_amount": parsed_data.get("total_amount", 0),
                "delivery_time_days": parsed_data.get("delivery_time_days"),
                "payment_terms": parsed_data.get("payment_terms"),
                "status": "External"
            })
        except Exception as e:
            continue
    
    if not all_quotations:
        raise HTTPException(status_code=404, detail="No quotations found")
    
    # Create Excel workbook
    wb = Workbook()
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4A5568", end_color="4A5568", fill_type="solid")
    highlight_fill = PatternFill(start_color="C6F6D5", end_color="C6F6D5", fill_type="solid")
    warning_fill = PatternFill(start_color="FED7D7", end_color="FED7D7", fill_type="solid")
    uploaded_fill = PatternFill(start_color="BEE3F8", end_color="BEE3F8", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Calculate statistics
    prices = [q["total_amount"] for q in all_quotations if q["total_amount"] > 0]
    lowest_price = min(prices) if prices else 0
    highest_price = max(prices) if prices else 0
    avg_price = sum(prices) / len(prices) if prices else 0
    
    # === Sheet 1: Summary ===
    ws_summary = wb.active
    ws_summary.title = "Combined Comparison"
    
    # Title
    ws_summary['A1'] = f"Combined Quotation Comparison - {request['title']}"
    ws_summary['A1'].font = Font(bold=True, size=16)
    ws_summary.merge_cells('A1:H1')
    
    ws_summary['A2'] = f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M')} UTC"
    ws_summary['A2'].font = Font(italic=True, color="666666")
    
    # Request Info
    ws_summary['A4'] = "Request Details"
    ws_summary['A4'].font = Font(bold=True, size=12)
    ws_summary['A5'] = f"Title: {request['title']}"
    ws_summary['A6'] = f"Budget: ${request.get('budget', 0):,.2f}"
    ws_summary['A7'] = f"Deadline: {request.get('deadline', 'N/A')}"
    
    # Statistics
    ws_summary['A9'] = "Price Analysis"
    ws_summary['A9'].font = Font(bold=True, size=12)
    
    stats = [
        ["Metric", "Value"],
        ["Total Quotations", len(all_quotations)],
        ["Vendor Submissions", len(db_quotations)],
        ["Uploaded PDFs", len(all_quotations) - len(db_quotations)],
        ["Lowest Price", f"${lowest_price:,.2f}" if lowest_price else "N/A"],
        ["Highest Price", f"${highest_price:,.2f}" if highest_price else "N/A"],
        ["Average Price", f"${avg_price:,.2f}" if avg_price else "N/A"],
        ["Potential Savings", f"${highest_price - lowest_price:,.2f}" if len(prices) >= 2 else "N/A"],
    ]
    
    for row_idx, row_data in enumerate(stats, start=10):
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            if row_idx == 10:
                cell.font = header_font
                cell.fill = header_fill
    
    # Vendor Comparison
    ws_summary['A20'] = "All Quotations Comparison"
    ws_summary['A20'].font = Font(bold=True, size=12)
    
    headers = ["Rank", "Vendor", "Source", "Total Amount", "Items", "Delivery", "Status", "Price vs Lowest"]
    for col_idx, header in enumerate(headers, start=1):
        cell = ws_summary.cell(row=21, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    sorted_quotations = sorted(all_quotations, key=lambda x: x["total_amount"] if x["total_amount"] > 0 else float('inf'))
    
    for row_idx, quot in enumerate(sorted_quotations, start=22):
        price_diff = ((quot["total_amount"] - lowest_price) / lowest_price * 100) if lowest_price > 0 and quot["total_amount"] > 0 else 0
        
        row_data = [
            row_idx - 21,
            quot["vendor_name"],
            quot["source"],
            quot["total_amount"],
            len(quot["items"]),
            f"{quot.get('delivery_time_days')}d" if quot.get('delivery_time_days') else "N/A",
            quot["status"],
            f"+{price_diff:.1f}%" if price_diff > 0 else ("Lowest" if quot["total_amount"] == lowest_price and lowest_price > 0 else "N/A")
        ]
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            
            if col_idx == 4 and isinstance(value, (int, float)):
                cell.number_format = '$#,##0.00'
            
            # Highlight
            if quot["total_amount"] == lowest_price and lowest_price > 0:
                cell.fill = highlight_fill
            elif quot["total_amount"] == highest_price and highest_price > 0 and len(prices) > 1:
                cell.fill = warning_fill
            elif "Uploaded" in quot["source"]:
                cell.fill = uploaded_fill
    
    for col in range(1, 9):
        ws_summary.column_dimensions[get_column_letter(col)].width = 18
    
    # === Sheet 2: All Items ===
    ws_items = wb.create_sheet("All Items Detail")
    
    item_headers = ["Vendor", "Source", "Item Description", "Quantity", "Unit Price", "Total Price"]
    for col_idx, header in enumerate(item_headers, start=1):
        cell = ws_items.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    row_idx = 2
    for quot in sorted_quotations:
        for item in quot["items"]:
            ws_items.cell(row=row_idx, column=1, value=quot["vendor_name"]).border = thin_border
            ws_items.cell(row=row_idx, column=2, value=quot["source"]).border = thin_border
            ws_items.cell(row=row_idx, column=3, value=item.get("product_name", "N/A")).border = thin_border
            ws_items.cell(row=row_idx, column=4, value=item.get("quantity", 1)).border = thin_border
            
            unit_cell = ws_items.cell(row=row_idx, column=5, value=item.get("unit_price", 0))
            unit_cell.number_format = '$#,##0.00'
            unit_cell.border = thin_border
            
            total_cell = ws_items.cell(row=row_idx, column=6, value=item.get("total_price", 0))
            total_cell.number_format = '$#,##0.00'
            total_cell.border = thin_border
            
            row_idx += 1
    
    ws_items.column_dimensions['A'].width = 25
    ws_items.column_dimensions['B'].width = 20
    ws_items.column_dimensions['C'].width = 45
    ws_items.column_dimensions['D'].width = 12
    ws_items.column_dimensions['E'].width = 15
    ws_items.column_dimensions['F'].width = 15
    
    # === Sheet 3: Item Price Matrix ===
    ws_matrix = wb.create_sheet("Price Matrix")
    
    all_items = {}
    for quot in sorted_quotations:
        for item in quot["items"]:
            item_name = item.get("product_name", "Unknown")
            if item_name not in all_items:
                all_items[item_name] = {}
            all_items[item_name][quot["vendor_name"]] = item.get("unit_price", 0)
    
    vendor_names = [q["vendor_name"] for q in sorted_quotations]
    
    matrix_headers = ["Item"] + vendor_names + ["Best Price", "Worst Price", "Savings"]
    for col_idx, header in enumerate(matrix_headers, start=1):
        cell = ws_matrix.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    row_idx = 2
    for item_name, vendor_prices in all_items.items():
        ws_matrix.cell(row=row_idx, column=1, value=item_name).border = thin_border
        
        prices_list = []
        for col_idx, vendor in enumerate(vendor_names, start=2):
            price = vendor_prices.get(vendor, 0)
            cell = ws_matrix.cell(row=row_idx, column=col_idx, value=price if price > 0 else "N/A")
            cell.border = thin_border
            if price > 0:
                cell.number_format = '$#,##0.00'
                prices_list.append(price)
        
        if prices_list:
            min_p = min(prices_list)
            max_p = max(prices_list)
            
            min_cell = ws_matrix.cell(row=row_idx, column=len(vendor_names) + 2, value=min_p)
            min_cell.number_format = '$#,##0.00'
            min_cell.fill = highlight_fill
            min_cell.border = thin_border
            
            max_cell = ws_matrix.cell(row=row_idx, column=len(vendor_names) + 3, value=max_p)
            max_cell.number_format = '$#,##0.00'
            max_cell.fill = warning_fill
            max_cell.border = thin_border
            
            savings_cell = ws_matrix.cell(row=row_idx, column=len(vendor_names) + 4, value=max_p - min_p)
            savings_cell.number_format = '$#,##0.00'
            savings_cell.border = thin_border
        
        row_idx += 1
    
    ws_matrix.column_dimensions['A'].width = 40
    for col in range(2, len(matrix_headers) + 1):
        ws_matrix.column_dimensions[get_column_letter(col)].width = 15
    
    # Save
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"combined_comparison_{request['title'][:30].replace(' ', '_')}_{datetime.utcnow().strftime('%Y%m%d')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )